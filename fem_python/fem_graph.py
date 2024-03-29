import matplotlib.pyplot as plt
import openpyxl
from fem_excel import *
import seaborn as sns

wk = Workbook()
model = model_size()


# P!=Falseで p=P, q=C, P==Falseで Cに番号, Mag=Trueでグラフに重さ表示, Dは分からん, coord="x" or "y" or "z"
def GraphMain(sheet, coord="z", P=False, C=1, D=False, Mag=False):
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    if type(sheet) is str:
        ws = wb[sheet]
    else:
        ws = wb.worksheets[sheet]

    a = 2
    yoko = []
    X = []
    Y = []
    Z = []
    mag = []

    if P != False:
        p = P
        q = C
        while not ws.cell(p, q).value is None:
            yoko.append(ws.cell(p, q).value)
            if coord == "x":
                X.append(ws.cell(p, q+1).value)
            elif coord == "z":
                Z.append(ws.cell(p, q+3).value)
            if Mag == True:
                mag.append(model["weight"])
            p += 1
    else:
        cal = 1+5*(C-1)
        while not ws.cell(a, cal).value is None:
            yoko.append(ws.cell(a, cal).value)
            X.append(ws.cell(a, cal+1).value)
            Y.append(ws.cell(a, cal+2).value)
            Z.append(ws.cell(a, cal+3).value)
            if Mag == True:
                mag.append(model["weight"])
            a += 1
        if D != False:
            X2 = []
            a = 2
            cal = 1+5*(D-1)
            while not ws.cell(a, cal).value is None:
                X2.append(ws.cell(a, cal+1).value)
                a += 1
    title = input("title:")
    fig = plt.figure(figsize=(10, 5))

    plt.title(title)
    if sheet == "rad":
        plt.xlabel("radius(mm)")
    elif sheet == "mag_num":
        plt.xlabel("num")
    elif sheet == "mesh":
        plt.xlabel("mesh(mm)")
    elif P != False:
        if sheet == 0 or sheet == 3:
            plt.xlabel("thick(mm)")
        elif sheet == 1 or sheet == 4:
            plt.xlabel("outer_radius(mm)")
        elif sheet == 2 or sheet == 5:
            plt.xlabel("inner_radius(mm)")
        else:
            plt.xlabel("height(mm)")
    else:
        plt.xlabel("height(mm)")
    plt.grid(True)

    if coord == "x":
        plt.ylabel("Fx(N)")
        plt.plot(yoko, X, color="blue", label="Fx")
        if D != False:
            plt.plot(yoko, X2, color="red", label="X:-3mm")
    elif coord == "z":
        plt.ylabel("Fz(N)")
        plt.plot(yoko, Z, color="blue", label="Fz")
    if Mag == True:
        plt.plot(yoko, mag, color="orange", label="Weight")
    plt.legend(loc="upper left", bbox_to_anchor=(1, 1))
    if sheet == "mag_num" or P != False:
        if title == "mesh(0.4~10mm)":
            yoko3 = []
            for i in yoko:
                if i >= 1:
                    yoko3.append(i)
            yoko3.append(0.5)
            plt.xticks(yoko3)
        elif title == "mesh(0.4~1mm)":
            j = 0
            yoko2 = []
            for i in yoko:
                if j == 0:
                    yoko2.append(i)
                    j = 1
                elif j == 1:
                    j = 0
            plt.xticks(yoko2)
        else:
            plt.xticks(yoko)
    else:
        j = 0
        yoko2 = []
        for i in yoko:
            if j == 0:
                yoko2.append(i)
                j = 1
            elif j == 1:
                j = 0
        plt.xticks(yoko2)
    fig.savefig(
        "documents/tanaka/git_files/fem_python/python_graph/"+title+".png")
    plt.clf()


# k,Cには番号　kからCまでをひとつのグラフにする, coord="x" or "y" or "z"
def GraphMain2(sheet, k=1, C=1, coord="x", square=False):
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    ws = wb[sheet]

    if square == True:
        fig = plt.figure(figsize=(10, 10))
    else:
        fig = plt.figure(figsize=(10, 5))
    colorlist = ["r", "g", "b", "c", "m", "y", "k"]
    plt.rcParams["font.size"] = 18
    title = input("title:")
    plt.title(title)
    yoko = []
    l = 2

    K = 1+5*(k-1)
    while not ws.cell(l, K).value is None:
        yoko.append(ws.cell(l, K).value)
        l += 1
    if sheet == "dis":
        plt.xlabel("height(mm)")
    elif sheet == "rad":
        plt.xlabel("radius(mm)")
    elif sheet == "mag_num":
        plt.xlabel("num")
    if coord == "z":
        plt.ylabel("Fz(N)")
    elif coord == "x":
        plt.ylabel("Fx(N)")
    plt.grid(True)
    if sheet == "mag_num":
        plt.xticks(yoko)
    else:
        j = 0
        yoko2 = []
        i = yoko[0]
        while i <= max(yoko):
            yoko2.append(i)
            i += 10
        plt.xticks(yoko2)
    j = 0
    while k <= C:
        i = 2
        Coord = []
        XYZ = {"x": 1, "y": 2, "z": 3}
        cal = 1 + XYZ[coord] + 5*(k-1)
        while not ws.cell(i, cal).value is None:
            Coord.append(ws.cell(i, cal).value)
            i += 1
        label_name = ws.cell(i+1, cal-XYZ[coord]).value
        plt.plot(yoko, Coord, color=colorlist[j], label=label_name)
        k += 1
        j += 1

    plt.legend(loc="lower right", bbox_to_anchor=(1, 0))
    plt.tight_layout()
    fig.savefig(
        "documents/tanaka/git_files/fem_python/python_graph/"+title+".png")


def GraphMain3(Data, Title=False, sheet="heatmap_z", NULLDATA=False, XZ=False):  # HeatmapMain用
    if NULLDATA == True:
        i = False
    else:
        i = True
    sns.heatmap(data=Data, annot=i, cmap="Blues", cbar_kws={
                "label": "height(mm)"}, vmin=0, vmax=70)
    if Title == True:
        title = input("title:")
    else:
        if sheet == "heatmap_z" or sheet == "heatmap_z2" and XZ == False:
            title = "Fz_stable_z"
        elif sheet == "heatmap_x":
            title = "Fx_stable_xy"
        else:
            title = "Fz_stable_xyz"
    plt.title(title)
    plt.savefig(
        "documents/tanaka/git_files/fem_python/python_graph/"+title+".png")
    plt.clf()


def GraphMain4(sheet, cals, coord="X"):  # calsにはリストで番号与える　番号のグラフを一つに表示　coord="x" or "y" "z"
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    ws = wb[sheet]

    fig = plt.figure(figsize=(10, 5))
    colorlist = ["r", "g", "b", "c", "m", "y", "k"]
    title = input("title:")
    plt.title(title)
    yoko = []
    l = 2

    K = 1+5*(cals[0]-1)
    while not ws.cell(l, K).value is None:
        yoko.append(ws.cell(l, K).value)
        l += 1
    if sheet == "dis":
        plt.xlabel("height(mm)")
    elif sheet == "rad":
        plt.xlabel("radius(mm)")
    elif sheet == "mag_num":
        plt.xlabel("num")
    if coord == "Z":
        plt.ylabel("Fz(N)")
    elif coord == "X":
        plt.ylabel("Fx(N)")
    plt.grid(True)
    if sheet == "mag_num":
        plt.xticks(yoko)
    else:
        j = 0
        yoko2 = []
        i = min(yoko)
        while i <= max(yoko):
            yoko2.append(i)
            i += 10
        plt.xticks(yoko2)
    j = 0
    for data in cals:
        i = 2
        Coord = []
        XYZ = {"x": 1, "y": 2, "z": 3}
        cal = 1 + XYZ[coord] + 5*(data-1)
        while not ws.cell(i, cal).value is None:
            Coord.append(ws.cell(i, cal).value)
            i += 1
        label_name = ws.cell(i+2, cal-XYZ[coord]).value
        plt.plot(yoko, Coord, color=colorlist[j], label=label_name)
        j += 1

    plt.legend(loc="lower right", bbox_to_anchor=(1, 0))
    fig.savefig("documents/tanaka/git_files/python_graph/"+title+".png")


# Title==Trueのときタイトル入力手動, NULLDATA==Trueのときデータ全て0
def HeatmapMain(Title=False, NULLDATA=False):
    E3_heatmap_move("data_z")
    GraphMain3(E3_heatmap_data("heatmap_z", NULLDATA),
               Title, "heatmap_z", NULLDATA)
    E3_heatmap_move("data_x")
    GraphMain3(E3_heatmap_data("heatmap_x", NULLDATA),
               Title, "heatmap_x", NULLDATA)
    E3_heatmap_xz()
    GraphMain3(E3_heatmap_data("heatmap_xz", NULLDATA),
               Title, "heatmap_xz", NULLDATA)


def HeatmapMain2(Title=False):  # sheetのz方向の復元力がはたらき始める範囲を表示
    E3_heatmap_remake("heatmap_z")
    GraphMain3(E3_heatmap_data("heatmap_z2"),
               Title, "heatmap_z2")
    E3_heatmap_remake("heatmap_xz")
    GraphMain3(E3_heatmap_data("heatmap_z2"),
               Title, "heatmap_z2", XZ=True)


def GraphMain5(sheet="data_z"):  # sheet内のデータを全てグラフ化
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    ws = wb[sheet]

    q = 1
    p = 2
    if sheet == "data_z":
        xz = 3
    elif sheet == "data_x":
        xz = 1
    while not ws.cell(p, q).value is None:
        fig = plt.figure(figsize=(10, 5))
        yoko = []
        Fz = []
        while not ws.cell(p, q).value is None:
            yoko.append(ws.cell(p, q).value)
            Fz.append(ws.cell(p, q+xz).value)
            p += 1
        title = ws.cell(p+1, q).value+","+ws.cell(p+2, q).value
        plt.title(title)
        plt.plot(yoko, Fz)
        fig.savefig(
            "documents/tanaka/git_files/fem_python/python_graph/"+title+".png")
        q += 5
        p = 2
        plt.clf()
