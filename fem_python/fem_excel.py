import openpyxl
from fem_MagSize import model_size
import sys
import numpy as np
import pandas as pd


class Count():
    StopCount = 0
    BackUpCount = True
    

class Workbook():
    book = "../fem_data/koma_sim3_py.xlsx"


check_model = model_size()
wk = Workbook()


def StopExit():
    Count.StopCount += 1
    if Count.StopCount > 100:
        sys.exit


def sheet_clear(ws):
    for row in ws:
        for cell in row:
            cell.value = None


def ExcelMain(ws, result, From, Interval):
    p = 1
    q = 1

    while not ws.cell(p, q).value is None:
        q += 5

    ws.cell(p, q).value = "z座標"
    ws.cell(p, q+1).value = "x"
    ws.cell(p, q+2).value = "y"
    ws.cell(p, q+3).value = "z"

    Q = q

    for i in result:
        p += 1
        ws.cell(p, Q).value = From
        Q += 1
        for j in i:
            ws.cell(p, Q).value = j
            Q += 1
        Q = q
        From += Interval

    p += 1

    return [p, Q]


def EThick(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[0]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "厚さ"+str(var["thick"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("厚さ"+str(var["thick"])+"mm"+"保存完了")


def EOutside(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[1]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "外半径"+str(var["outside"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("外半径"+str(var["outside"])+"mm"+"保存完了")


def EInside(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[2]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "内半径"+str(var["inside"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("内半径"+str(var["inside"])+"mm"+"保存完了")


def EL_Thick(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[3]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "厚さ"+str(var["L_thick"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("大厚さ"+str(var["L_thick"])+"mm"+"保存完了")


def EL_Outside(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[4]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "外半径"+str(var["L_outside"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("大外半径"+str(var["L_outside"])+"mm"+"保存完了")


def EL_Inside(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[5]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "内半径"+str(var["L_inside"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("大内半径"+str(var["L_inside"])+"mm"+"保存完了")


def ECoord_x(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[6]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "x座標"+str(var["coord_x"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("x座標"+str(var["coord_x"])+"mm"+"保存完了")


def ECoord_y(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[7]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "y座標"+str(var["coord_y"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("y座標"+str(var["coord_y"])+"mm"+"保存完了")


def Emeshsize(result, From, Interval, var):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[8]

    P = ExcelMain(ws, result, From, Interval)

    ws.cell(P[0], P[1]).value = "メッシュ"+str(var["meshsize"])+"mm"

    wb.save("koma_sim_py.xlsx")
    print("メッシュ"+str(var["meshsize"])+"mm"+"保存完了")


def E_sub(result, comment):
    wb = openpyxl.load_workbook("koma_sim_py.xlsx")
    ws = wb.worksheets[9]

    p = 1
    q = 1

    while not ws.cell(p, q).value is None:
        p += 4

    xyz = ["x", "y", "z"]
    ws.cell(p, q).value = comment
    j = 0
    for i in result:
        ws.cell(p+1, q).value = xyz[j]
        ws.cell(p+2, q).value = i
        q += 1
        j += 1

    wb.save("koma_sim_py.xlsx")
    print(comment+"保存完了")


def E2_org2(sheet, C, m):  # 上下のデータを交互に入れる
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]
    while C <= m:
        Q = 1+5*(C-1)
        q = Q
        p = 1
        list0 = []
        list1 = []
        list2 = []
        list3 = []
        f = 0
        while f <= 4:
            list0.append(ws.cell(p, q).value)
            q += 1
            f += 1
        p += 1
        q = Q
        while not ws.cell(p, q).value is None:
            list_cal = []
            while not ws.cell(p, q).value is None:
                list_cal.append(ws.cell(p, q).value)
                q += 1
            list1.append(list_cal)
            p += 1
            q = Q
        p += 3
        while not ws.cell(p, q).value is None:
            list_cal = []
            while not ws.cell(p, q).value is None:
                list_cal.append(ws.cell(p, q).value)
                q += 1
            list2.append(list_cal)
            p += 1
            q = Q
        p += 1
        while not ws.cell(p, q).value is None:
            list3.append(ws.cell(p, q).value)
            p += 1
        p = 1
        ws.delete_cols(Q, 5)
        ws.insert_cols(Q, 5)
        for i in list0:
            ws.cell(p, q).value = i
            q += 1
        p = 2
        q = Q
        for i in list1:
            for i_cal in i:
                ws.cell(p, q).value = i_cal
                q += 1
            p += 2
            q = Q
        p = 3
        for i in list2:
            for i_cal in i:
                ws.cell(p, q).value = i_cal
                q += 1
            p += 2
            q = Q
        p += 1
        for i in list3:
            ws.cell(p, q).value = i
            p += 1
        C += 1
    wb.save(wk.book)
    print(sheet+"整理完了")


def E2_org3():  # 行を上下反転
    wb = wb = openpyxl.load_workbook("koma_sim2_py.xlsx")
    ws = wb["mesh"]

    p = 2
    q = 4
    MoveData = []
    while not ws.cell(p, q).value is None:
        MoveData.insert(0, ws.cell(p, q).value)
        p += 1

    p = 2
    q = 7
    for i in MoveData:
        ws.cell(p, q).value = i
        p += 1

    wb.save("koma_sim2_py.xlsx")


def E3_hanntenn(sheet="heatmap_z", P=3, Q=2):  # E3_heatmap_z修正用
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]
    p = P
    q = Q
    data = []
    while not ws.cell(p, q).value is None:
        data.insert(0, [])
        while not ws.cell(p, q).value is None:
            data[0].append(ws.cell(p, q).value)
            q += 1
        p += 1
        q = Q
    p = P
    for i in data:
        for j in i:
            ws.cell(p, q).value = j
            q += 1
        p += 1
        q = Q

    wb.save(wk.book)


"""
↑リング型フェライト磁石の解析用＆使わないやつ
↓複数のネオジム磁石の解析用
"""


def E2_org(sheet, WS=False):  # 全体を右に5ずらす
    if WS == False:
        wb = openpyxl.load_workbook(wk.book)
        ws = wb[sheet]
    else:
        ws = WS
    ws.insert_cols(1, 5)
    StopExit()
    if WS == False:
        wb.save(wk.book)
    print(sheet+"整理完了")


def E2_comment(sheet, comment):  # セルにコメントを記録
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]

    p = 1
    while not ws.cell(p, 1).value is None:
        p += 1
    p += 1
    while not ws.cell(p, 1).value is None:
        p += 1
    ws.cell(p, 1).value = comment
    wb.save(wk.book)
    E_BackUp(comment, True)
    print(comment+"完了")


def E2_check(sheet, model):  # 解析時のモデルの大きさを"check"シートに記録
    wb = openpyxl.load_workbook(wk.book)
    ws = wb["check"]
    ws2 = wb[sheet]
    q = 1
    p = 1

    if ws.cell(p, q).value is None:
        ws.cell(p, q).value = sheet

    while not ws.cell(p, q).value is None:
        p += 1
    if ws2.cell(p, q).value is None:
        p += 1
        ws.cell(p, q).value = ws2.cell(p, q+5).value
        E2_org("check", ws)
        wb.save(wk.book)
        E2_check(sheet, model)
        return
    else:
        ws.cell(p, q).value = ws2.cell(p, q).value
        q += 1
        for i in model:
            if model[i] != check_model[i]:
                ws.cell(p, q).value = i+":"+str(model[i])
                q += 1

    wb.save(wk.book)


def ExcelMain2(result, subj, var):  # 結果をセルに記録
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[var]

    p = 1
    q = 1

    if ws.cell(p, q).value is None:
        ws.cell(p, q).value = "mm"
        ws.cell(p, q+1).value = "x"
        ws.cell(p, q+2).value = "y"
        ws.cell(p, q+3).value = "z"
        if subj == "mag_num":
            ws.cell(p, q).value = "num"

    while not ws.cell(p, q).value is None:
        p += 1

    ws.cell(p, q).value = subj
    for i in result:
        q += 1
        ws.cell(p, q).value = i

    wb.save(wk.book)
    print(var + str(subj)+"mm"+"保存完了")


def E2main(result, model, var):  # これを実行すればよい
    ExcelMain2(result, model[var], var)
    E_BackUp(result, model[var])
    E2_check(var, model)
    E2_number(var)


def E2_number(sheet, load_book=False):  # データの列に番号を振る
    if sheet == "all":
        wb = openpyxl.load_workbook(wk.book)
        i = 0
        while i < len(wb.worksheets):
            ws = wb.worksheets[i]
            C = 1
            while not ws.cell(1, 1+5*(C-1)).value is None:
                q = 5+5*(C-1)
                ws.cell(1, q).value = C
                C += 1
            if load_book != True:
                wb.save(wk.book)
            i += 1

    elif load_book == True:
        ws = sheet
    else:
        wb = openpyxl.load_workbook(wk.book)
        ws = wb[sheet]

    C = 1
    while not ws.cell(1, 1+5*(C-1)).value is None:
        q = 5+5*(C-1)
        ws.cell(1, q).value = C
        C += 1
    if load_book != True:
        wb.save(wk.book)


def E_BackUp(result, subj):  # バックアップ用ブックに記録
    wb = openpyxl.load_workbook(
        "C:/users/skmgr/documents/tanaka_back_up/excel_data/data_backup.xlsx")
    ws = wb.worksheets[0]
    p = 1
    q = 1

    if Count.BackUpCount == True and subj != True:
        ws.insert_cols(1, 5)
        Count.BackUpCount = False

    while not ws.cell(p, q).value is None:
        p += 1
    if subj == True:
        p += 1
        while not ws.cell(p, q).value is None:
            p += 1
        ws.cell(p, q).value = result
        Count.BackUpCount = True
    else:
        ws.cell(p, q).value = subj
        for i in result:
            q += 1
            ws.cell(p, q).value = i
    E2_number(ws, True)
    p = 1
    q = 10
    while not ws.cell(p, q).value is None:
        q += 5
    q += -5
    if ws.cell(p, q).value >= 100:
        wb.create_sheet(index=0, title="data_backup"+str(len(wb.worksheets)+1))
    wb.save("C:/users/skmgr/documents/tanaka_back_up/excel_data/data_backup.xlsx")


def E2_delete_sheet(sheet):  # シートを消去
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]
    wb.remove(ws)
    wb.save(wk.book)
    print(sheet+"削除完了")


def E3_heatmap_data(sheet):  # ヒートマップ用二次元配列データを返す
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]

    p = 3
    q = 3

    Move_MainData = []
    while not ws.cell(p, q).value is None:
        while not ws.cell(p, q).value is None:
            Move_MainData.append(ws.cell(p, q).value)
            p += 1
        mag_num = p-3
        p = 3
        q += 1
    rad = q-3
    MainData = pd.DataFrame(data=Move_MainData, columns=["Fz(N)"])

    i = 0
    Move_MagNumData = []
    while i < rad:
        p = 3
        q = 2
        while not ws.cell(p, q).value is None:
            Move_MagNumData.append(ws.cell(p, q).value)
            p += 1
        i += 1
    MainData["num"] = Move_MagNumData

    i = 0
    p = 2
    q = 3
    Move_RadData = []
    while not ws.cell(p, q).value is None:
        while i < mag_num:
            Move_RadData.append(ws.cell(p, q).value)
            i += 1
        i = 0
        q += 1
    MainData["rad(mm)"] = Move_RadData

    MainData_pivot = pd.pivot_table(
        data=MainData, values="Fz(N)", columns="rad(mm)", index="num")
    MainData_pivot.sort_index(ascending=False, inplace=True)

    return MainData_pivot


def E3_heatmap_move(sheet="data_z"):  # 解析データからヒートマップ用データを作成
    wb = openpyxl.load_workbook(wk.book)
    ws = wb[sheet]
    if sheet == "data_z":
        ws2 = wb["heatmap_z"]
    elif sheet == "data_x":
        ws2 = wb["heatmap_x"]

    sheet_clear(ws2)

    p = 2
    q_F = 4
    q_x = 2
    q_dis = 1

    weight = 0.0611  # 重さを変える(単位:N)

    L = 13
    while not ws.cell(p, q_F).value is None:
        find = False
        while not ws.cell(p, q_F).value is None:
            if sheet == "data_z":
                if p >= 3 and p <= 7 and find == False:
                    if ws.cell(p, q_F).value > weight and ws.cell(p+1, q_F).value < weight:
                        ws.cell(L, q_dis).value = ws.cell(p, q_dis).value
                        find = True
            elif sheet == "data_x":
                if find == False:
                    if ws.cell(p, q_x).value >= 0:
                        ws.cell(L, q_dis).value = ws.cell(p-1, q_dis).value
                        if isinstance(ws.cell(L, q_dis).value, str):
                            ws.cell(L, q_dis).value = 0
                        find = True
            p += 1
        if find == False:
            if sheet == "data_x":
                ws.cell(L, q_dis).value = ws.cell(p-1, q_dis).value
            else:
                ws.cell(L, q_dis).value = 0
        q_dis += 5
        q_F += 5
        q_x += 5
        p = 2

    m = 13
    n = 1
    i = 0
    move_data_z = [[]]
    Mag = ws.cell(m-2, n).value
    while not ws.cell(m, n).value is None:
        if Mag != ws.cell(m-2, n).value:
            Mag = ws.cell(m-2, n).value
            move_data_z.append([])
            i += 1
        move_data_z[i].insert(0, ws.cell(m, n).value)
        n += 5
    p = 3
    q = 3

    for i in move_data_z:
        for j in i:
            ws2.cell(p, q).value = j
            q += 1
        q = 3
        p += 1
    rad = 30
    mag_num = 0
    ws2.cell(1, 3).value = "rad"
    ws2.cell(3, 1).value = "mag_num"
    P = 3
    Q = 3
    while not ws2.cell(P, Q).value is None:
        ws2.cell(P-1, Q).value = rad
        rad += 1
        Q += 1
    Q = 3
    while not ws2.cell(P, Q).value is None:
        mag_num += 1
        P += 1
    P = 3
    while not ws2.cell(P, Q).value is None:
        ws2.cell(P, Q-1).value = mag_num
        mag_num += -1
        P += 1

    wb.save(wk.book)


def E3_heatmap_xz():  # X方向も考慮したヒートマップ用データを作成
    wb = openpyxl.load_workbook(wk.book)
    ws = wb["heatmap_z"]
    ws_x = wb["heatmap_xz"]
    ws_XData = wb["data_x"]

    sheet_clear(ws_x)

    p = 3
    q = 3
    mag_num=ws.cell(3,2).value
    while not ws.cell(p, q).value is None:
        while not ws.cell(p, q).value is None:
            if ws.cell(p, q).value != 0:
                num = (q-2)+11*(2+mag_num-p)
                j = 2
                k = 1+5*(num-1)
                while not ws_XData.cell(j, k).value is None:
                    if ws_XData.cell(j, k).value == ws.cell(p, q).value:
                        if ws_XData.cell(j, k+1).value >= 0:
                            ws_x.cell(p, q).value = ws.cell(p, q).value
                        else:
                            ws_x.cell(p, q).value = 0
                    j += 1
            else:
                ws_x.cell(p, q).value = 0
            q += 1
        p += 1
        q = 3
    rad = 30
    mag_num = 0
    ws_x.cell(1, 3).value = "rad"
    ws_x.cell(3, 1).value = "mag_num"
    P = 3
    Q = 3
    while not ws_x.cell(P, Q).value is None:
        ws_x.cell(P-1, Q).value = rad
        rad += 1
        Q += 1
    Q = 3
    while not ws.cell(P, Q).value is None:
        mag_num += 1
        P += 1
    P = 3
    while not ws_x.cell(P, Q).value is None:
        ws_x.cell(P, Q-1).value = mag_num
        mag_num += -1
        P += 1

    wb.save(wk.book)
