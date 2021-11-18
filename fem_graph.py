import matplotlib.pyplot as plt
import openpyxl
from fem_excel import Workbook

wk = Workbook()


def GraphMain(sheet, coord="z", P=False, C=1, D=False, Mag=False):
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    if type(sheet) is str:
        ws = wb[sheet]
    else:
        ws = wb.worksheets[sheet]

    a = 2
    yoko = []
    X = []
    Y = []
    Z = []
    mag = []

    if P != False:
        p = P
        q = C
        while not ws.cell(p, q).value is None:
            yoko.append(ws.cell(p, q).value)
            if coord == "x":
                X.append(ws.cell(p, q+1).value)
            elif coord == "z":
                Z.append(ws.cell(p, q+1).value)
            p += 1
    else:
        cal = 1+5*(C-1)
        while not ws.cell(a, cal).value is None:
            yoko.append(ws.cell(a, cal).value)
            X.append(ws.cell(a, cal+1).value)
            Y.append(ws.cell(a, cal+2).value)
            Z.append(ws.cell(a, cal+3).value)
            if Mag == True:
                mag.append(ws.cell(a, cal+4).value)
            a += 1
        if D != False:
            X2 = []
            a = 2
            cal = 1+5*(D-1)
            while not ws.cell(a, cal).value is None:
                X2.append(ws.cell(a, cal+1).value)
                a += 1
    title = input("title:")
    fig = plt.figure(figsize=(10, 5))

    plt.title(title)
    if sheet == "rad":
        plt.xlabel("radius(mm)")
    elif sheet == "mag_num":
        plt.xlabel("num")
    elif sheet == "mesh":
        plt.xlabel("mesh(mm)")
    elif P != False:
        if sheet == 0 or sheet == 3:
            plt.xlabel("thick(mm)")
        elif sheet == 1 or sheet == 4:
            plt.xlabel("outer_radius(mm)")
        elif sheet == 2 or sheet == 5:
            plt.xlabel("inner_radius(mm)")
    else:
        plt.xlabel("height(mm)")
    plt.grid(True)

    if coord == "x":
        plt.ylabel("Fx(N)")
        plt.plot(yoko, X, color="blue", label="X:3mm")
        if D != False:
            plt.plot(yoko, X2, color="red", label="X:-3mm")
    elif coord == "z":
        plt.ylabel("Fz(N)")
        plt.plot(yoko, Z, color="blue", label="Fz")
    if Mag == True:
        plt.plot(yoko, mag, color="orange", label="Weight")
    plt.legend(loc="upper left", bbox_to_anchor=(1, 1))
    if sheet == "mag_num" or P != False:
        if title == "mesh(0.4~10mm)":
            yoko3 = []
            for i in yoko:
                if i >= 1:
                    yoko3.append(i)
            yoko3.append(0.5)
            plt.xticks(yoko3)
        elif title == "mesh(0.4~1mm)":
            j = 0
            yoko2 = []
            for i in yoko:
                if j == 0:
                    yoko2.append(i)
                    j = 1
                elif j == 1:
                    j = 0
            plt.xticks(yoko2)
        else:
            plt.xlabel(yoko)
    else:
        j = 0
        yoko2 = []
        for i in yoko:
            if j == 0:
                yoko2.append(i)
                j = 1
            elif j == 1:
                j = 0
        plt.xticks(yoko2)
    fig.savefig(
        "python_graph/"+title+".png")


def GraphMain2(sheet, k=1, C=1, coord="X"):
    wb = openpyxl.load_workbook(wk.book, data_only=True)
    ws = wb[sheet]

    fig = plt.figure(figsize=(10, 5))
    colorlist = ["r", "g", "b", "c", "m", "y", "k"]
    title = input("title:")
    plt.title(title)
    yoko = []
    l = 2

    K = 1+5*(k-1)
    while not ws.cell(l, K).value is None:
        yoko.append(ws.cell(l, K).value)
        l += 1
    if sheet == "dis":
        plt.xlabel("height(mm)")
    elif sheet == "rad":
        plt.xlabel("radius(mm)")
    elif sheet == "mag_num":
        plt.xlabel("num")
    if coord == "Z":
        plt.ylabel("Fz(N)")
    elif coord == "X":
        plt.ylabel("Fx(N)")
    plt.grid(True)
    if sheet == "mag_num":
        plt.xticks(yoko)
    else:
        j = 0
        yoko2 = []
        i = min(yoko)
        while i <= max(yoko):
            yoko2.append(i)
            i += 10
        plt.xticks(yoko2)
    j = 0
    while k <= C:
        i = 2
        Coord = []
        XYZ = {"X": 1, "Y": 2, "Z": 3}
        cal = 1 + XYZ[coord] + 5*(k-1)
        while not ws.cell(i, cal).value is None:
            Coord.append(ws.cell(i, cal).value)
            i += 1
        label_name = ws.cell(i+1, cal-XYZ[coord]).value
        plt.plot(yoko, Coord, color=colorlist[j], label=label_name)
        k += 1
        j += 1

    plt.legend(loc="lower right", bbox_to_anchor=(1, 0))
    fig.savefig("python_graph/"+title+".png")


def GraphMain3(sheet, rad_ticks, num_ticks, data):
    plt.figure(figsize=(10, 10))
    i = rad_ticks[0]
    yoko = []
    while i <= rad_ticks[1]:
        yoko.append(i)
        i += 10
    plt.xticks(yoko)
    tate = []
    j = num_ticks[0]
    while j <= num_ticks[1]:
        tate.append(j)
        j += 1
    plt.yticks(tate)
    plt.imshow(data, cmap="gray", vmin=0, vmax=255, interpolation="nearest")
    plt.colorbar()
    plt.show()
